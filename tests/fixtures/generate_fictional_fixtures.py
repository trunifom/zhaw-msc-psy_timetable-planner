"""
Regenerates the large-scale and scenario-specific fictional test fixtures
in tests/fixtures/ from the user's real, local ZHAW export files in
data/real/ (git-ignored, never committed - see README.md "Datenschutz").

Why this script exists: docs/planung/KONZEPT-passerelle-zusatzmodule.md's
Phase 1/2 work was verified against the real HS26 Bachelor/Master
catalogs, which uncovered real-world patterns (e.g. the "undistinguished
parallel offerings" pattern - see app._has_undistinguished_parallel_
offerings) that a hand-written fixture is unlikely to reproduce
faithfully or at realistic scale. Rather than hand-crafting more fixture
rows one at a time, this script fictionalizes the *actual* real catalogs
end to end, so the generated fixtures inherit the real scheduling
diversity (dates, times, Modulart mix, parallel-group shapes) while never
containing any real person's name or real ZHAW curriculum content.

Fictionalization rules (see the standing project convention - real
provided files must never be committed unmodified, only fictionalized
first):
  - KEPT as-is: SG, Semester, Datum, von, bis, Modulart, the Pruefung
    flag/Tag column. None of this is personally identifying or specific
    curriculum content - it's scheduling metadata, and preserving it
    exactly is the whole point of these fixtures (realistic dates/times/
    module-type mix, not just realistic *shape*).
  - FICTIONALIZED via a deterministic (same real value -> same fictional
    value on every run, so regenerating is stable/diffable) mapping:
    Lehrperson (-> a Muster/Beispiel/Demo/Platzhalter/Vorlage/Fiktiv-style
    name, matching this project's established fixture convention),
    Modul-Nr/Kurs-Nr (-> a synthetic code using letter-prefixes that don't
    exist in either real catalog, keeping the original numeric suffix so
    a real module's *family* of course components - e.g. "AFE1".."AFE4" -
    still maps to a consistent fictional family), and Anlassbezeichnung
    (-> a generated "Adjektiv + Nomen" title, with any real "/Gruppe A",
    "- Halbklasse 1", "/ Pruefung" etc. suffix preserved verbatim since
    that structural pattern - not its wording - is what tests exercise).

This script also (re)writes a handful of small, hand-engineered scenario
fixture pairs (conflict detection, a Modul-Nr collision between the two
lists) that are NOT derived from real data at all - see section 6 below -
because a real catalog, even fictionalized, can't reliably guarantee a
specific cross-catalog condition a test needs to assert on (e.g. "these
two rows must overlap in time"). Those are always regenerated; the two
full-scale fictionalized catalogs (section on Scenario 1 in main()) only
run when data/real/ is present locally.

Run manually (not part of the pytest suite, not part of CI) whenever a
new/updated real export lands in data/real/ and the fixtures need
refreshing:
    python tests/fixtures/generate_fictional_fixtures.py

Requires data/real/Vorlesungsdaten BSc HS26_P.xlsx and
data/real/Vorlesungsdaten MSc HS26.xlsx to exist locally - the script
exits early with a clear message if they don't (e.g. on a machine that
never had the real files, or in CI), since it has nothing real to
fictionalize from in that case. The *output* fixtures it writes ARE
committed (they're fictional); this script itself is also committed
(it contains no real data - only logic and a pool of fictional
names/titles) purely so the provenance of the generated fixtures stays
transparent and reproducible for future sessions.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
REAL_BSC_PATH = REPO_ROOT / "data" / "real" / "Vorlesungsdaten BSc HS26_P.xlsx"
REAL_MSC_PATH = REPO_ROOT / "data" / "real" / "Vorlesungsdaten MSc HS26.xlsx"
FIXTURES_DIR = REPO_ROOT / "tests" / "fixtures"


# ==========================================================================
# 1. FICTIONAL NAME POOLS
# ==========================================================================
# Same "obviously invented" surname convention already used throughout
# this project's existing fixtures (Muster/Beispiel/Demo/Platzhalter/
# Vorlage/Fiktiv + a plausible first name) - never a real-sounding full
# name, so nobody could mistake a fixture row for a real person.
_SURNAMES = [
    "Muster", "Beispiel", "Demo", "Platzhalter", "Vorlage", "Fiktiv",
    "Testfall", "Musterfrau", "Musterhaft", "Probeweise",
]
_FIRST_NAMES = [
    "Anna", "Kevin", "Rea", "Fabienne", "Jonas", "Noemi", "Timo", "Sina",
    "Selina", "Elias", "Priska", "Yann", "Aline", "Nora", "Levin", "Mia",
    "Colin", "Lara", "Dario", "Zoe", "Nico", "Salome", "Finn", "Ines",
    "Timon", "Alessia", "Julian", "Melina", "Silvan", "Anja",
]
_FICTIONAL_FULL_NAMES = [f"{s} {f}" for s in _SURNAMES for f in _FIRST_NAMES]  # 300 combos

# Fictional letter-prefixes for Modul-Nr/Kurs-Nr - deliberately built from
# letter combinations that don't occur as a first-letter-pair in either
# real catalog (verified below in build_modul_prefix_map's assertion), so
# a fictional code can never be mistaken for - or collide with - a real
# ZHAW module code.
_FICTIONAL_PREFIXES = [
    "QX", "QY", "QZ", "XJ", "XL", "XQ", "XT", "XW", "XY", "XZ",
    "YJ", "YL", "YQ", "YT", "YW", "YX", "YZ", "ZJ", "ZL", "ZQ",
    "ZT", "ZW", "ZX", "ZY", "JQ", "JT", "JW", "JX", "JY", "JZ",
]

# "Adjektiv"/"Nomen" pools combined into fictional course titles (e.g.
# "Grundlagen der Testtheorie") - psychology-adjacent-sounding but
# invented, matching the style already established in
# vorlesungsverzeichnis_passerelle_fiktiv.xlsx's hand-written rows.
_TITLE_PREFIXES = [
    "Grundlagen der", "Angewandte", "Einfuehrung in", "Methoden der",
    "Vertiefung in", "Theorien der", "Praxis der", "Modelle der",
    "Konzepte der", "Perspektiven der",
]
_TITLE_TOPICS = [
    "Testtheorie", "Entwicklungspsychologie", "Sozialpsychologie",
    "Kognitionspsychologie", "Diagnostik", "Beratung", "Intervention",
    "Forschungsmethodik", "Statistik", "Gespraechsfuehrung",
    "Verhaltensanalyse", "Lernpsychologie", "Wahrnehmungspsychologie",
    "Emotionsregulation", "Gruppendynamik", "Persoenlichkeitspsychologie",
    "Klinische Psychologie", "Arbeitspsychologie", "Gesundheitspsychologie",
    "Neuropsychologie", "Entscheidungsverhalten", "Motivationspsychologie",
    "Stressbewaeltigung", "Konfliktmanagement", "Teamdynamik",
    "Verhaltenstherapie", "Bindungstheorie", "Resilienzforschung",
    "Praeventionsarbeit", "Fallanalyse", "Testkonstruktion",
]
# 10 prefixes x 30 topics = 300 combos - comfortable headroom over the
# ~197 distinct base titles the real combined catalogs currently need
# (see _assign_deterministically's ValueError if this pool is ever too
# small for a future, larger real export).
_FICTIONAL_TITLES = [f"{p} {t}" for p in _TITLE_PREFIXES for t in _TITLE_TOPICS]


def _assign_deterministically(real_values: list[str], pool: list[str]) -> dict[str, str]:
    """
    Map each distinct real value to one pool entry, assigned by the real
    value's position in *sorted* order - deterministic across runs
    (unlike e.g. Python's randomized string hashing) so regenerating the
    fixtures from the same real input always produces the same fictional
    output. Falls back to a numbered suffix if the pool is smaller than
    the number of distinct real values (shouldn't happen with the pool
    sizes above, but fails loudly rather than silently colliding two real
    values onto the same fictional one if it ever did).
    """
    ordered = sorted(set(real_values))
    if len(ordered) > len(pool):
        raise ValueError(
            f"Fictional pool too small: {len(ordered)} distinct real values, only {len(pool)} pool entries."
        )
    return {real: pool[i] for i, real in enumerate(ordered)}


# ==========================================================================
# 2. SUFFIX-PRESERVING TITLE SPLITTING
# ==========================================================================
# A deliberately-simplified, local re-implementation of the suffix
# vocabulary app._split_course_variant() recognizes (kept separate from
# that function rather than importing it, since importing app.py here
# would pull in Streamlit just to generate test fixtures) - only used to
# decide which trailing words of a real course title are a *structural*
# marker (kept verbatim, since that's what parallel-group/exam detection
# tests exercise) versus the base title itself (fictionalized).
_SUFFIX_PATTERN = re.compile(
    r"\s*[/\-]\s*("
    r"ganzklasse|"
    r"gruppe\s+[a-z0-9][a-z0-9\s&+\-]*|"
    r"tk\d+\s+gruppe\s+[a-z](?:\s*&\s*gruppe\s*[a-z])?|"
    r"durchf(?:u|ue|ü)hrung\s*\d+|"
    r"pr(?:u|ue|uef|uf|ü)fung"
    r")\s*$",
    re.IGNORECASE,
)


def split_title_suffix(raw_title: str) -> tuple[str, str]:
    """Peel known variant/exam suffixes off the end of a title, one at a
    time (a title can carry more than one, e.g. ".../Gruppe A/Pruefung"),
    returning (base_title_without_suffixes, suffixes_in_original_order)."""
    name = str(raw_title or "").strip()
    suffixes: list[str] = []
    while True:
        match = _SUFFIX_PATTERN.search(name)
        if not match:
            break
        suffixes.insert(0, match.group(0))
        name = name[: match.start()].strip()
    return name, "".join(suffixes)


# ==========================================================================
# 3. LOADING THE REAL FILES
# ==========================================================================

def load_real_bsc() -> pd.DataFrame:
    df = pd.read_excel(REAL_BSC_PATH, sheet_name="Sheet", header=6)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_real_msc() -> pd.DataFrame:
    xls = pd.ExcelFile(REAL_MSC_PATH)
    df = pd.read_excel(REAL_MSC_PATH, sheet_name=xls.sheet_names[0], header=2)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ==========================================================================
# 4. FICTIONALIZATION
# ==========================================================================

def build_modul_prefix_map(all_modul_nrs: list[str]) -> dict[str, str]:
    """Map each real Modul-Nr's *letter prefix* (e.g. "AFE" from "AFE1")
    to one fictional prefix, so a real module family like AFE1..AFE4 maps
    to a consistent fictional family (e.g. QX1..QX4) instead of each row
    getting an unrelated-looking code."""
    real_prefixes = sorted({re.match(r"^([A-Za-z]+)", str(v)).group(1) for v in all_modul_nrs if pd.notna(v)})
    assert not (set(real_prefixes) & set(_FICTIONAL_PREFIXES)), "fictional prefix pool collides with a real one"
    return _assign_deterministically(real_prefixes, _FICTIONAL_PREFIXES)


def fictionalize_modul_nr(real_value: str, prefix_map: dict[str, str]) -> str:
    match = re.match(r"^([A-Za-z]+)(\d*)$", str(real_value).strip())
    if not match:
        return str(real_value)
    prefix, digits = match.group(1), match.group(2)
    return prefix_map[prefix] + digits


def fictionalize_kurs_nr(real_value: str, modul_nr_fictional: str) -> str:
    """Kurs-Nr in both real catalogs is "{Modul-Nr}-{n}" - reuse the
    already-fictionalized Modul-Nr and keep only the "-{n}" suffix from
    the real value, so e.g. real "AFE1-2" with fictional Modul-Nr "QX1"
    becomes "QX1-2", not a value unrelated to its own Modul-Nr column."""
    match = re.search(r"-(\S+)$", str(real_value).strip())
    return f"{modul_nr_fictional}-{match.group(1)}" if match else modul_nr_fictional


def fictionalize_dataframe(
    df: pd.DataFrame,
    name_map: dict[str, str],
    prefix_map: dict[str, str],
    title_map: dict[str, str],
) -> pd.DataFrame:
    """
    Return a copy of `df` with Lehrperson/Modul-Nr./Kurs-Nr./
    Anlassbezeichnung fictionalized via the given maps; every other
    column (SG, Semester, Datum, von, bis, Pruefung, Modulart, Tag if
    present) is left untouched. See the module docstring for the
    rationale.

    The three maps are passed in (built once, from the *combined*
    Bachelor+Master real data - see main()) rather than built fresh per
    call: building them separately per file would assign fictional
    values by each file's own sorted-index position, which can - and, in
    an earlier version of this script, silently did - map two entirely
    different real Modul-Nr prefixes (one from each catalog) onto the
    same fictional prefix purely by coincidence of sort order. That would
    fabricate a Modul-Nr collision between the two catalogs that doesn't
    exist in the real data (see docs/planung/KONZEPT-passerelle-
    zusatzmodule.md section 8, risk 2, which relies on the real catalogs'
    namespaces being disjoint) - caught by tests/test_zusatzmodule_
    fixtures.py::test_full_scale_catalogs_have_no_modul_nr_overlap.
    """
    out = df.copy()

    out["Lehrperson"] = out["Lehrperson"].map(lambda v: name_map.get(v, v) if pd.notna(v) else v)

    modul_nr_map = {
        real: fictionalize_modul_nr(real, prefix_map)
        for real in out["Modul-Nr."].dropna().unique()
    }
    out["Modul-Nr."] = out["Modul-Nr."].map(lambda v: modul_nr_map.get(v, v) if pd.notna(v) else v)

    out["Kurs-Nr."] = [
        fictionalize_kurs_nr(real_kurs, modul_nr_map.get(real_modul, real_modul))
        if pd.notna(real_kurs) and pd.notna(real_modul)
        else real_kurs
        for real_kurs, real_modul in zip(df["Kurs-Nr."], df["Modul-Nr."])
    ]

    # Titles are mapped per BASE title (suffix stripped first, see
    # split_title_suffix) so every row of "Kurs X - Gruppe A"/"Kurs X -
    # Gruppe B"/etc. shares the same fictional base "Kurs X" got mapped
    # to, and the real "- Gruppe A"/"- Gruppe B" structural distinction
    # (what the parallel-groups tests care about) survives untouched.
    def _map_title(raw: object) -> object:
        if pd.isna(raw):
            return raw
        base, suffix = split_title_suffix(str(raw))
        return f"{title_map[base]}{suffix}" if base in title_map else raw

    out["Anlassbezeichnung"] = out["Anlassbezeichnung"].map(_map_title)

    return out


# ==========================================================================
# 5. WRITING FIXTURE FILES (banner-row structure matching the real exports)
# ==========================================================================

def write_fixture(
    df: pd.DataFrame,
    columns: list[str],
    out_path: Path,
    title_banner: str,
    note: str,
    blank_column_after: str | None = None,
) -> None:
    """
    Write one fixture .xlsx with the same "title banner row + note row +
    real header row + data" shape every existing fixture in this project
    uses (see docs/TESTING-README.md) - this is what exercises
    data_loader._try_reheader_from_rows()'s header-detection logic, not a
    convenience of this script.

    `blank_column_after`: when set (e.g. "SG" for the Master-shaped
    layout), inserts one empty column right after that column - mirrors
    the real Master export's blank "Semester"-less second column, so a
    fixture generated for that shape also exercises _normalize_columns()'s
    "drop unnamed columns" behaviour, same as the real file does.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vorlesungsverzeichnis"

    header = list(columns)
    if blank_column_after is not None:
        idx = header.index(blank_column_after) + 1
        header.insert(idx, "")

    last_col = len(header)
    ws.cell(row=1, column=1, value=title_banner)
    ws.cell(row=1, column=last_col - 1, value="Stand:")
    ws.cell(row=1, column=last_col, value="27.08.2026")
    ws.cell(row=2, column=1, value=note)
    for col_idx, col_name in enumerate(header, start=1):
        ws.cell(row=3, column=col_idx, value=col_name)

    row_idx = 4
    for _, row in df.iterrows():
        col_idx = 1
        for col_name in header:
            if col_name == "":
                col_idx += 1
                continue
            value = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(value) else value)
            if col_name == "Datum" and pd.notna(value):
                cell.number_format = "yyyy-mm-dd h:mm:ss"
            col_idx += 1
        row_idx += 1

    wb.save(out_path)
    print(f"wrote {out_path.relative_to(REPO_ROOT)} ({len(df)} rows)")


BSC_COLUMNS = ["SG", "Semester", "Lehrperson", "Datum", "von", "bis", "Modul-Nr.", "Kurs-Nr.", "Anlassbezeichnung", "Prüfung", "Modulart"]
MSC_COLUMNS = ["SG", "Lehrperson", "Tag", "Datum", "von", "bis", "Modul-Nr.", "Kurs-Nr.", "Anlassbezeichnung", "Prüfung", "Modulart"]


# ==========================================================================
# 6. HAND-ENGINEERED SCENARIO FIXTURES (not derived from real data)
# ==========================================================================
# Unlike the two "vollstaendig" fixtures above, these are small, entirely
# invented DataFrames - not fictionalized real rows - built to deterministically
# exercise one specific behaviour end-to-end through the real loader/merge/
# conflict pipeline, which a real (even if fictionalized) catalog can't
# reliably guarantee (e.g. "these two specific rows must overlap in time").
# They still use write_fixture() for the same banner-row file shape as
# every other fixture in this project.

def _scenario_df(columns: list[str], rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=columns)


def write_conflict_scenario() -> None:
    """
    A main-list (Master-shaped) + Zusatzmodule (Bachelor-shaped) fixture
    pair with one deliberately overlapping row pair (same date, overlapping
    time) and one deliberately non-overlapping pair - so a test can assert
    find_time_conflicts() catches exactly the former and not the latter,
    driven through the *real* load -> merge -> conflict pipeline (not
    hand-built ZHAWModule objects, which tests/test_scheduler.py already
    covers at the unit level).
    """
    main_df = _scenario_df(
        MSC_COLUMNS,
        [
            {
                "SG": "MSc", "Lehrperson": "Konflikt Haupt", "Tag": "Di",
                "Datum": pd.Timestamp("2026-09-15"), "von": "10:15", "bis": "12:00",
                "Modul-Nr.": "KFH1", "Kurs-Nr.": "KFH1-1", "Anlassbezeichnung": "Szenario Hauptmodul",
                "Prüfung": None, "Modulart": "Pflichtmodul 1. Studienjahr",
            },
            {
                "SG": "MSc", "Lehrperson": "Konflikt Haupt", "Tag": "Mi",
                "Datum": pd.Timestamp("2026-09-16"), "von": "08:15", "bis": "10:00",
                "Modul-Nr.": "KFH2", "Kurs-Nr.": "KFH2-1", "Anlassbezeichnung": "Szenario Hauptmodul ohne Konflikt",
                "Prüfung": None, "Modulart": "Pflichtmodul 1. Studienjahr",
            },
        ],
    )
    zusatz_df = _scenario_df(
        BSC_COLUMNS,
        [
            {
                # Same date as KFH1, 10:15-11:45 overlaps KFH1's 10:15-12:00 -
                # a genuine conflict. No "Tag" column here on purpose (BSc
                # shape), so this also exercises weekday-from-datum
                # derivation landing on the same weekday as the main row's
                # explicit "Di" - both must resolve to the same weekday for
                # find_time_conflicts() to even compare them.
                "SG": "BSc", "Semester": "1. Semester VZ", "Lehrperson": "Konflikt Zusatz",
                "Datum": pd.Timestamp("2026-09-15"), "von": "10:15", "bis": "11:45",
                "Modul-Nr.": "KFZ1", "Kurs-Nr.": "KFZ1-1", "Anlassbezeichnung": "Szenario Zusatzmodul",
                "Prüfung": None, "Modulart": "Pflichtmodul",
            },
            {
                # Different date entirely from KFH2 - must NOT be flagged.
                "SG": "BSc", "Semester": "1. Semester VZ", "Lehrperson": "Konflikt Zusatz",
                "Datum": pd.Timestamp("2026-09-17"), "von": "08:15", "bis": "10:00",
                "Modul-Nr.": "KFZ2", "Kurs-Nr.": "KFZ2-1", "Anlassbezeichnung": "Szenario Zusatzmodul ohne Konflikt",
                "Prüfung": None, "Modulart": "Pflichtmodul",
            },
        ],
    )
    write_fixture(
        main_df, MSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_konflikt_hauptliste_fiktiv.xlsx",
        "Vorlesungsverzeichnis Master HS 2026 (TESTDATEN, FIKTIV, KONFLIKT-SZENARIO)",
        "Von Hand erstelltes Testszenario (nicht aus echten Daten abgeleitet) fuer Konflikterkennung "
        "zwischen Haupt- und Zusatzliste - siehe tests/fixtures/generate_fictional_fixtures.py.",
        blank_column_after="SG",
    )
    write_fixture(
        zusatz_df, BSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_konflikt_zusatzliste_fiktiv.xlsx",
        "Vorlesungsverzeichnis Bachelor HS 2026 (TESTDATEN, FIKTIV, KONFLIKT-SZENARIO)",
        "Von Hand erstelltes Testszenario (nicht aus echten Daten abgeleitet) fuer Konflikterkennung "
        "zwischen Haupt- und Zusatzliste - siehe tests/fixtures/generate_fictional_fixtures.py.",
    )


def write_modul_nr_collision_scenario() -> None:
    """
    A main-list + Zusatzmodule fixture pair that both use the exact same
    Modul-Nr ("KOL1") - see docs/planung/KONZEPT-passerelle-zusatzmodule.md
    section 8, risk 2. Not something either real catalog actually does
    (verified: zero overlap between the real 70 BSc / 21 MSc Modul-Nr
    codes), but worth a dedicated regression fixture since the app-level
    grouping behaviour for this case was clarified (not "fixed") during
    the Phase 1 re-review - a future change to the grouping logic should
    have this concrete case to check against.
    """
    main_df = _scenario_df(
        MSC_COLUMNS,
        [{
            "SG": "MSc", "Lehrperson": "Kollision Haupt", "Tag": "Do",
            "Datum": pd.Timestamp("2026-10-01"), "von": "08:15", "bis": "10:00",
            "Modul-Nr.": "KOL1", "Kurs-Nr.": "KOL1-1", "Anlassbezeichnung": "Kollisions-Szenario Hauptliste",
            "Prüfung": None, "Modulart": "Pflichtmodul 1. Studienjahr",
        }],
    )
    zusatz_df = _scenario_df(
        BSC_COLUMNS,
        [{
            "SG": "BSc", "Semester": "1. Semester VZ", "Lehrperson": "Kollision Zusatz",
            "Datum": pd.Timestamp("2026-10-02"), "von": "10:15", "bis": "12:00",
            "Modul-Nr.": "KOL1", "Kurs-Nr.": "KOL1-1", "Anlassbezeichnung": "Kollisions-Szenario Zusatzliste",
            "Prüfung": None, "Modulart": "Pflichtmodul",
        }],
    )
    write_fixture(
        main_df, MSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_modulnr_kollision_hauptliste_fiktiv.xlsx",
        "Vorlesungsverzeichnis Master HS 2026 (TESTDATEN, FIKTIV, MODUL-NR-KOLLISION)",
        "Von Hand erstelltes Testszenario (nicht aus echten Daten abgeleitet) fuer eine Modul-Nr., "
        "die zufaellig in Haupt- und Zusatzliste identisch ist - siehe tests/fixtures/generate_fictional_fixtures.py.",
        blank_column_after="SG",
    )
    write_fixture(
        zusatz_df, BSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_modulnr_kollision_zusatzliste_fiktiv.xlsx",
        "Vorlesungsverzeichnis Bachelor HS 2026 (TESTDATEN, FIKTIV, MODUL-NR-KOLLISION)",
        "Von Hand erstelltes Testszenario (nicht aus echten Daten abgeleitet) fuer eine Modul-Nr., "
        "die zufaellig in Haupt- und Zusatzliste identisch ist - siehe tests/fixtures/generate_fictional_fixtures.py.",
    )


def main() -> None:
    # These two don't need data/real/ at all - always (re)written first so
    # they stay available even on a machine without the real exports.
    write_conflict_scenario()
    write_modul_nr_collision_scenario()

    if not REAL_BSC_PATH.exists() or not REAL_MSC_PATH.exists():
        print(
            "Real files not found under data/real/ - nothing to fictionalize from for the "
            "full-scale fixtures. This is expected on a machine that never had the real "
            "exports (e.g. CI); the already-generated fixtures in tests/fixtures/ are unaffected.",
            file=sys.stderr,
        )
        return

    real_bsc = load_real_bsc()
    real_msc = load_real_msc()

    # Built once from the COMBINED real data (not per-file - see
    # fictionalize_dataframe()'s docstring for why that matters) so a
    # real property like "Bachelor and Master Modul-Nr namespaces never
    # collide" carries over into the fictional output instead of being
    # accidentally fabricated or destroyed by independent sort orders.
    combined_names = pd.concat([real_bsc["Lehrperson"], real_msc["Lehrperson"]])
    name_map = _assign_deterministically(combined_names.dropna().unique().tolist(), _FICTIONAL_FULL_NAMES)

    combined_modul_nrs = pd.concat([real_bsc["Modul-Nr."], real_msc["Modul-Nr."]])
    prefix_map = build_modul_prefix_map(combined_modul_nrs.dropna().unique().tolist())

    combined_titles = pd.concat([real_bsc["Anlassbezeichnung"], real_msc["Anlassbezeichnung"]])
    combined_base_titles = [split_title_suffix(v)[0] for v in combined_titles.dropna()]
    title_map = _assign_deterministically(combined_base_titles, _FICTIONAL_TITLES)

    fictional_bsc = fictionalize_dataframe(real_bsc, name_map, prefix_map, title_map)
    fictional_msc = fictionalize_dataframe(real_msc, name_map, prefix_map, title_map)

    # --- Scenario 1: full-scale fictionalized catalogs ------------------
    # The whole point of "umfangreiche Testdaten" - realistic scale (full
    # real row count) and realistic diversity (every Modulart/Semester/
    # parallel-group pattern the real catalogs actually contain), not a
    # hand-picked small excerpt.
    write_fixture(
        fictional_bsc,
        BSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_bsc_vollstaendig_fiktiv.xlsx",
        "Vorlesungsverzeichnis Bachelor HS 2026 (TESTDATEN, FIKTIV, VOLLSTAENDIG)",
        "Vollstaendig fiktionalisierter Testdatensatz (Struktur/Umfang des echten Bachelor-Vorlesungsverzeichnisses). "
        "Keine echten Personen oder Module - siehe tests/fixtures/generate_fictional_fixtures.py.",
    )
    write_fixture(
        fictional_msc,
        MSC_COLUMNS,
        FIXTURES_DIR / "vorlesungsverzeichnis_msc_vollstaendig_fiktiv.xlsx",
        "Vorlesungsverzeichnis Master HS 2026 (TESTDATEN, FIKTIV, VOLLSTAENDIG)",
        "Vollstaendig fiktionalisierter Testdatensatz (Struktur/Umfang des echten Master-Vorlesungsverzeichnisses). "
        "Keine echten Personen oder Module - siehe tests/fixtures/generate_fictional_fixtures.py.",
        blank_column_after="SG",
    )


if __name__ == "__main__":
    main()
